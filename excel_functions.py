# function file for Excel operations

# excel_functions.py

from openpyxl import load_workbook


class Suman_Excel_Functions:

    def __init__(self, excel_file_name, sheet_name):
        self.file = excel_file_name
        self.sheet = sheet_name

    # Get row count of the excel sheet
    def Row_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_row)

    # Get column count of the excel sheet
    def Column_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_column)

    # Read data from excel file
    def Read_Data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.cell(row=row_number, column=column_number).value)
    
    # Write data to the excel file
    def Write_Data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)
